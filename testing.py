from datetime import date, datetime
from openpyxl import workbook, cell, load_workbook
from os import listdir

run = True  # Crea una variable para que el script siga corriendo siempre que el usuario quiera
ganancias = input("Ubicacion del archivo DDJJ Ganancias: ")


def libroF(hoja, mes, año, libro):
    global netogravado
    global nogravado
    global regEsp
    global importetotal
    global fila
    # Recorre la columna A en busca de la fila donde estan los valores, gracias a la variable totalmes
    for celda in hoja['A']:
        if celda.value == mes:
            fila = celda.row
            if libro == 1:
                # Busca los valores del neto gravado, el IVA, y el total en base a la fila donde encontro a la variable totalmes
                netogravado = hoja[f"D{fila}"].value
                importetotal = hoja[f"F{fila}"].value
                return netogravado and importetotal
            elif libro == 2:
                # Cambia las celdas en las que se encuentran los valores deseados segun el libro donde se busque, debido a que esto ocasiono un error en el programa
                netogravado = hoja[f"E{fila}"].value
                nogravado = hoja[f"B{fila}"].value
                regEsp = hoja[f"C{fila}"].value
                importetotal = hoja[f"G{fila}"].value
                return netogravado, nogravado, regEsp, importetotal
        elif celda.value == f"TOTALES AL 29/02/{año}":
            fila = celda.row
            if libro == 1:
                netogravado = hoja[f"D{fila}"].value
                importetotal = hoja[f"F{fila}"].value
                return netogravado and importetotal
            elif libro == 2:
                netogravado = hoja[f"E{fila}"].value
                nogravado = hoja[f"B{fila}"].value
                regEsp = hoja[f"C{fila}"].value
                importetotal = hoja[f"G{fila}"].value
                return netogravado, nogravado, regEsp, importetotal


def alicuota(hoja, hoja2, mes, año, libro):
    global valoresIVA
    valoresIVA = []
    for celda in hoja['B']:
        if celda.value == 21:
            valoresIVA.append("IVA 21%")
        elif celda.value == 10.5:
            valoresIVA.append("IVA 10.5%")
        elif celda.value == 27:
            valoresIVA.append("IVA 27%")
    for celda2 in hoja2['A']:
        if celda2.value == datetime(int(año), mes, 1, 0, 0, 0):
            lista = str(set(valoresIVA))
            if libro == 1:
                hoja2.cell(row=celda2.row, column=9).value = lista
            elif libro == 2:
                hoja2.cell(row=celda2.row, column=7).value = lista


def ventas(hojaV, mes):  # Crea una funcion que busca dentro del archivo DDJJ Ganancias la fila y la columna en la cual debera ubicar los datos
    global fila1
    global columna
    for celda in hojaV['A']:
        if celda.value == mes:
            fila1 = celda.row
            for celda2 in hojaV['C']:
                if celda2.value == "$":  # La diferencia entre "$" y "gravado" es que en algunos archivos aparece una, y en otros aparecen ambas. Por eso es una solucion global.
                    fila2 = celda2.row
                    if hojaV.cell(row=fila2, column=5).value == "gravado":
                        columna = 5
                        return columna
                        break
                    else:
                        columna = celda2.column
                        return columna
                        break
                elif celda2.value == "gravadas":
                    columna = celda2.column
                    return columna
            return fila1


gastos = ["BCO.CREDICOOP - ARGENCARD", "BANCO CREDICOOP - CABAL", "EDENOR SA", "TELEFONICA",
          "AMERICAN EXPRESS SA", "REDGUARD SA", "BANCO CREDICOOP", "POSNET SRL", "CULLIGAN ARG SA", "EUROTIME SA", "posnet srl"]


def compras(hojaC, hojaG, mes):
    global valor1
    global valor2
    global gastoTotal
    global rowC
    global columnC
    for celda in hojaC['D']:
        for gasto in gastos:  # Busca cada uno de los gastos definidos en la lista de la linea 27
            if celda.value == gasto:  # Compara el valor de la celda con los gastos
                # Al tener uno o dos valores (ya que hay importes con un iva de 10,5% y otros con 21%)
                row = celda.row + 1
                # Busca los valores de las 2 filas siguientes a donde se encuentra el concepto gasto
                row2 = celda.row + 2
                # Como los valores siempre estan en la columna A (o column=1), toma esos valores
                valor1 = hojaC.cell(row=row, column=1).value
                valor2 = hojaC.cell(row=row2, column=1).value
                if valor2 == None:  # Evita el error por NoneType, ya que si encuentra que el segundo valor es nulo, solo suma el valor1
                    gastoTotal = + valor1
                elif type(valor1) and type(valor2) == float:
                    gastoTotal = + valor1 + valor2
    for celda2 in hojaG['A']:
        if celda2.value == mes:
            rowC = celda2.row
            columnasPosibles = [hojaG['G6'], hojaG['H6']]
            for columna in columnasPosibles:
                if columna.value == "Gastos":
                    columnC = columna.column
                    return gastoTotal, rowC and columnC
                elif columna.value == "gastos":
                    columnC = columna.column
                    return gastoTotal, rowC and columnC

def mes(libro, type):
    global mesL
    if type == 1:
        num1 = int(libro[-6])
        try:
            num2 = int(libro[-7])
            mesL = f"{num2}{num1}"
            return str(mesL)   
        except:
            mesL = num1
            return str(mesL)
    elif type == 2:
        num1 = int(libro[-6])
        try:
            num2 = int(libro[-7])
            mesL = f"{num2}{num1}"
            return str(mesL)
        except:
            mesL = num1
            return str(mesL)

año = str(input('Año de los libros (ejemplo: 2020): '))
pathV = input("Inserte el directorio donde se encuentren los libros de ventas: ")
pathC = input("Inserte el directorio donde se encuentren los libros de compras: ")
librosV = listdir(pathV)
librosC = listdir(pathC)

total_mes = {
    "1": f"TOTALES AL 31/01/{año}",
    "2": f"TOTALES AL 28/02/{año}",
    "3": f"TOTALES AL 31/03/{año}",
    "4": f"TOTALES AL 30/04/{año}",
    "5": f"TOTALES AL 31/05/{año}",
    "6": f"TOTALES AL 30/06/{año}",
    "7": f"TOTALES AL 31/07/{año}",
    "8": f"TOTALES AL 31/08/{año}",
    "9": f"TOTALES AL 30/09/{año}",
    "10": f"TOTALES AL 31/10/{año}",
    "11": f"TOTALES AL 30/11/{año}",
    "12": f"TOTALES AL 31/12/{año}"
}
nombre_mes = {
    1: "ENERO",
    2: "FEBRERO",
    3: "MARZO",
    4: "ABRIL",
    5: "MAYO",
    6: "JUNIO",
    7: "JULIO",
    8: "AGOSTO",
    9: "SEPTIEMBRE",
    10: "OCTUBRE",
    11: "NOVIEMBRE",
    12: "DICIEMBRE",
}
while run == True:
    for libro in librosV:
        total = total_mes[mes(libro, 1)]
        nombre = nombre_mes[int(mes(libro, 1))]
        archivo = load_workbook(f"{pathV}{libro}")
        archivo2 = load_workbook(ganancias)
        hoja = archivo.active
        libroF(hoja, total, año, 1)
        hoja2 = archivo2['VENTAS']
        ventas(hoja2, nombre)
        hoja2.cell(row=fila1, column=columna).value = netogravado
        alicuota(hoja, hoja2, int(mes(libro, 1)), año, 1)
        archivo2.save(ganancias)
    for libro in librosC:
        total = total_mes[mes(libro, 2)]
        nombre = nombre_mes[int(mes(libro, 2))]
        archivoC = load_workbook(f"{pathC}{libro}")
        archivo2 = load_workbook(ganancias)
        hojaC = archivoC.active
        hoja2 = archivo2['COMPRAS']
        compras(hojaC, hoja2, nombre)
        print(rowC, columnC)
        libroF(hojaC, total, año, 2)
        hoja2.cell(row=rowC, column= (columnC)).value = gastoTotal
        hoja2.cell(row=rowC, column= (columnC - 2)).value = (netogravado - gastoTotal)
        alicuota(hojaC, hoja2, int(mes(libro, 2)), año, 2)
        archivo2.save(ganancias)       
    break    